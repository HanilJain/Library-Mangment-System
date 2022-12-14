#!/usr/bin/python3
from openpyxl import workbook,load_workbook 
from openpyxl.utils import get_column_letter
import openpyxl

def populate(color,books,path):
    list_book_name = list(books.keys())
    wb  = load_workbook(path)
    ws = wb.active 
    try :
        for c1, c2 ,c3,c4,c5 in ws[ws.dimensions]:
                print(color.RED + color.UNDERLINE + "*"*50)
                print(color.BLUE+"Book Name : " + c1.value)
                print(color.RED+'Book Author : ' +  c2.value)
                print(color.lightgreen+"Book ISBN : " + c3.value)
                print(color.CYAN+"Book Type : " + c4.value)
                print(color.GREEN+"Book Status : " + c5.value)
                print(color.RED + color.UNDERLINE + "*"*50+'\n')



    except IndexError:
        pass




